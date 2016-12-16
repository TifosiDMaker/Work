from openpyxl import Workbook
from openpyxl.styles import PatternFill,Fill
from openpyxl.styles.colors import YELLOW
from openpyxl import styles
from openpyxl.compat import range
import os
from os.path import join

wb = Workbook()
ws = wb.active
#dest_dir = input('请输入外部审校文件所在路径.\n>')
#dest_dir = dest_dir.replace("\\","/")


#i = 1
#for root, dirs, files in os.walk(dest_dir):
   #for name in files:
       #ws['B' + str(i)].value = name
       #i = int(i)
       #i += 1
fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
ws['A1'] = 'dff'
ws['A1'].fill
wb.save('d:/Tifosi/test.xlsx')
